# environments/view.py
"""Project views."""

import csv
import logging
import os
from datetime import datetime

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import PermissionDenied
from django.core.paginator import Paginator
from django.db.models import Count, Q
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse, reverse_lazy
from django.views.generic import (
    CreateView,
    DeleteView,
    DetailView,
    ListView,
    TemplateView,
    UpdateView,
    View,
)

from chat.models import ChatSession, Message
from chat.services.rag_service import SYSTEM_PROMPT, RAGService
from core.llm import resolve_model_name
from core.mixins import (
    ProjectAccessMixin,
    ProjectOwnerRequiredMixin,
    ProjectTabMixin,
)
from core.token_budget import compute_budget, get_context_window
from documents.models import Document
from documents.services.document_processor import DocumentProcessor
from ifc_processor.models import IFCEntity, IFCFile, IFCSpatialElement
from ifc_processor.services.processor import IFCProcessingService
from ifc_processor.services.schedule_service import (
    IFC_TYPE_CATEGORIES,
    DoorWindowScheduleService,
    GenericScheduleService,
)
from ifc_processor.templatetags.ifc_filters import get_unit_type_for_quantity, smart_round

from .models import Project, ProjectMembership, ProjectRole
from .services import (
    LastOwnerRemovalBlocked,
    OwnerDemotionBlocked,
    ProjectAccessError,
    ProjectAccessService,
)

logger = logging.getLogger(__name__)


def _fmt_ctx(tokens: int) -> str:
    """Format a context window token count as a compact label, e.g. '8k', '32k'."""
    return f"{tokens // 1024}k" if tokens >= 1024 else str(tokens)


class ProjectListView(LoginRequiredMixin, ListView):
    """List all projects the user is a member of, regardless of tier."""

    model = Project
    template_name = "environments/project_list.html"
    context_object_name = "projects"

    def get_queryset(self):
        """Projects with a ProjectMembership for the current user."""
        return (
            ProjectAccessService.accessible_projects(self.request.user)
            .filter(is_archived=False)
            .select_related("owner")
            .annotate(
                ifc_count=Count("ifc_files", distinct=True),
                document_count=Count("documents", distinct=True),
                conflict_count=Count(
                    "conflicts", filter=Q(conflicts__status="open"), distinct=True
                ),
                issue_count=Count(
                    "ifc_files__data_issues",
                    filter=Q(ifc_files__data_issues__status="open"),
                    distinct=True,
                ),
            )
            .order_by("-updated_at")
        )


class ProjectCreateView(LoginRequiredMixin, CreateView):
    """Create a new project and bootstrap its OWNER membership."""

    model = Project
    template_name = "environments/project_form.html"
    fields = ["name", "description"]
    success_url = reverse_lazy("projects:list")

    def form_valid(self, form):
        form.instance.owner = self.request.user
        response = super().form_valid(form)
        # Keep Project.owner and the OWNER membership row in lockstep from
        # creation onward. Without this bootstrap, the creator holds the FK
        # but has no membership row — every access check refuses them.
        ProjectAccessService.bootstrap_owner_membership(self.object)
        return response


class ProjectDetailView(ProjectAccessMixin, DetailView):
    """Project detail - redirects to Ask tab by default."""

    model = Project
    template_name = "environments/project_detail.html"
    context_object_name = "project"

    def get_object(self):
        return self.get_project()

    def get(self, request, *args, **kwargs):
        """Redirect to Ask tab by default."""
        return redirect("projects:ask", pk=self.kwargs["pk"])


# --- Project CRUD ---
class ProjectUpdateView(ProjectOwnerRequiredMixin, UpdateView):
    """Edit project details (name, description)."""

    model = Project
    template_name = "environments/project_form.html"
    fields = ["name", "description"]
    context_object_name = "project"

    def get_success_url(self):
        # Redirect back to the project dashboard (Ask tab)
        return reverse("projects:ask", kwargs={"pk": self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = f"Edit {self.object.name}"
        return context


class ProjectDeleteView(ProjectOwnerRequiredMixin, DeleteView):
    """Delete a project and all associated data."""

    model = Project
    template_name = "environments/confirm_delete.html"
    success_url = reverse_lazy("projects:list")
    context_object_name = "object"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = f"Delete Project: {self.object.name}"
        context["message"] = (
            "Are you sure you want to delete this project? All IFC files, documents, and chat history will be permanently lost."
        )
        return context


class AskView(ProjectTabMixin, TemplateView):
    """Ask tab — multi-session chat against IFC/documents."""

    active_tab = "ask"

    def _get_or_create_session(self, project, user):
        """Return the most recent Ask session, or create one."""
        session = (
            ChatSession.objects.filter(project=project, user=user, mode=ChatSession.Mode.ASK)
            .order_by("-updated_at")
            .first()
        )
        if not session:
            session = ChatSession.objects.create(
                project=project,
                user=user,
                mode=ChatSession.Mode.ASK,
                title="New conversation",
            )
        return session

    def _resolve_session(self, project, user):
        """Resolve session from URL kwarg or fallback to most recent."""
        session_id = self.kwargs.get("session_id")
        if session_id:
            return get_object_or_404(
                ChatSession,
                pk=session_id,
                project=project,
                user=user,
                mode=ChatSession.Mode.ASK,
            )
        return self._get_or_create_session(project, user)

    def get(self, request, *args, **kwargs):
        """Redirect bare /ask/ to the most recent session URL."""
        if "session_id" not in self.kwargs:
            project = self.get_project()
            session = self._get_or_create_session(project, request.user)
            return redirect("projects:ask_session", pk=project.pk, session_id=session.pk)
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        project = self.get_project()
        session = self._resolve_session(project, self.request.user)

        context["session"] = session
        context["active_session_id"] = session.pk
        messages_qs = session.messages.select_related().order_by("created_at")
        context["messages"] = messages_qs

        if messages_qs.exists():
            model_name = resolve_model_name(self.request.user)
            history = [{"role": m.role, "content": m.content} for m in messages_qs]
            budget = compute_budget(model_name, system=SYSTEM_PROMPT, conversation_history=history)
            context["utilization_pct"] = budget.utilization_pct
            context["context_window_label"] = _fmt_ctx(budget.model_context_window)

        return context

    def post(self, request, *args, **kwargs):
        project = self.get_project()
        user_text = request.POST.get("message", "").strip()
        scope = request.POST.get("scope", "auto")

        # "New Chat" button — create session and redirect
        if request.POST.get("action") == "new_session":
            session = ChatSession.objects.create(
                project=project,
                user=request.user,
                mode=ChatSession.Mode.ASK,
                title="New conversation",
            )
            return redirect("projects:ask_session", pk=project.pk, session_id=session.pk)

        if not user_text:
            return redirect("projects:ask", pk=project.pk)

        session = self._resolve_session(project, request.user)

        # Save user message
        Message.objects.create(
            session=session,
            role=Message.Role.USER,
            content=user_text,
        )

        # Generate AI response
        utilization_pct = 0.0
        try:
            rag = RAGService(user=request.user)
            answer_text, context_items, utilization_pct = rag.generate_answer(
                project,
                session,
                user_text,
                scope=scope,
            )
            Message.objects.create(
                session=session,
                role=Message.Role.ASSISTANT,
                content=answer_text,
                retrieved_context=rag._serialize_context(context_items, scope=scope),
            )
        except Exception as e:
            logger.exception("RAG generation failed for session %s: %s", session.pk, e)
            Message.objects.create(
                session=session,
                role=Message.Role.ASSISTANT,
                content="⚠️ I encountered an error processing your question. "
                "Please check that Ollama is running and try again.",
            )

        # Auto-title on first exchange
        if not session.title or session.title == "New conversation":
            session.generate_title()

        # HTMX partial response
        if request.headers.get("HX-Request"):
            updated_messages = session.messages.select_related().order_by("created_at")
            model_name = resolve_model_name(request.user)
            return render(
                request,
                "environments/components/chat_message_list.html",
                {
                    "messages": updated_messages,
                    "user": request.user,
                    "utilization_pct": utilization_pct,
                    "context_window_label": _fmt_ctx(get_context_window(model_name)),
                },
            )

        return redirect("projects:ask_session", pk=project.pk, session_id=session.pk)


class AskMessagesView(ProjectTabMixin, View):
    """
    Return the rendered message list partial for a session.

    Called via htmx.ajax() after the WebSocket pipeline emits {type: "done"},
    so the browser can swap in server-rendered messages (including markdown
    and source badges) without a full page reload.
    """

    active_tab = "ask"

    def get(self, request, pk, session_id, *args, **kwargs):
        project = self.get_project()
        session = get_object_or_404(
            ChatSession,
            pk=session_id,
            project=project,
            user=request.user,
            mode=ChatSession.Mode.ASK,
        )
        messages_qs = (
            session.messages.select_related()
            .exclude(is_compaction_summary=True)
            .order_by("created_at")
        )
        model_name = resolve_model_name(request.user)
        history = [{"role": m.role, "content": m.content} for m in messages_qs]
        budget = compute_budget(model_name, system=SYSTEM_PROMPT, conversation_history=history)
        return render(
            request,
            "environments/components/chat_message_list.html",
            {
                "messages": messages_qs,
                "user": request.user,
                "utilization_pct": budget.utilization_pct,
                "context_window_label": _fmt_ctx(budget.model_context_window),
            },
        )


class DeleteSessionView(ProjectAccessMixin, View):
    """Delete a chat session and redirect to Ask tab."""

    def post(self, request, pk, session_id):
        project = self.get_project()
        session = get_object_or_404(
            ChatSession,
            pk=session_id,
            project=project,
            user=request.user,
        )
        session.delete()
        logger.info("Deleted chat session %s for project %s", session_id, pk)
        return redirect("projects:ask", pk=project.pk)


class RenameSessionView(ProjectAccessMixin, View):
    """Rename a chat session title."""

    def post(self, request, pk, session_id):
        project = self.get_project()
        session = get_object_or_404(
            ChatSession,
            pk=session_id,
            project=project,
            user=request.user,
        )
        title = request.POST.get("title", "").strip()
        if title:
            session.title = title[:255]
            session.save(update_fields=["title"])
            logger.info("Renamed session %s to '%s'", session_id, session.title)

        return JsonResponse({"ok": True, "title": session.title})


class IFCFileUpdateView(ProjectAccessMixin, UpdateView):
    """Replace an IFC file and re-parse it."""

    model = IFCFile
    fields = ["file"]  # ONLY file, name auto-updates
    template_name = "environments/file_form.html"

    def get_project(self):
        return self.get_object().project

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = f"Replace File: {self.object.name}"
        context["help_text"] = (
            "Uploading a new file will replace the current model and automatically re-parse all entities."
        )
        return context

    def form_valid(self, form):
        # 1. Save the new file
        self.object = form.save(commit=False)
        self.object.name = self.object.file.name  # Auto-sync name
        self.object.save()

        # 2. Run Pipeline (Hash -> Parse -> Embed) using the Service
        processor = IFCProcessingService(self.object)
        success = processor.run_pipeline()

        # 3. Store result for the next page
        self.request.session["processing_result"] = {
            "success": success,
            "filename": self.object.name,
            "entity_count": self.object.entity_count,
            "error": self.object.error_message,
            "type": "IFC Model",
        }

        return redirect("projects:file_processed", pk=self.object.project.pk)


class IFCReparseView(LoginRequiredMixin, View):
    """Confirmation page and trigger for reparsing an already-uploaded IFC file.

    Reparse re-runs the processing pipeline, wipes extracted entities, and
    may churn the Git history. OWNER only.
    """

    def _get_ifc(self, request, pk) -> IFCFile:
        ifc_file = get_object_or_404(IFCFile.objects.select_related("project"), pk=pk)
        if not ProjectAccessService.can_delete(request.user, ifc_file.project):
            raise PermissionDenied("Only the project owner can reparse IFC files.")
        return ifc_file

    def get(self, request: HttpRequest, pk) -> HttpResponse:
        ifc_file = self._get_ifc(request, pk)
        return render(
            request,
            "environments/ifc_reparse_confirm.html",
            {"ifc_file": ifc_file, "project": ifc_file.project},
        )

    def post(self, request: HttpRequest, pk) -> HttpResponse:
        ifc_file = self._get_ifc(request, pk)
        processor = IFCProcessingService(ifc_file)
        success = processor.run_pipeline()
        request.session["processing_result"] = {
            "success": success,
            "filename": ifc_file.name,
            "entity_count": ifc_file.entity_count,
            "error": ifc_file.error_message,
            "type": "IFC Model",
        }
        return redirect("projects:file_processed", pk=ifc_file.project.pk)


class IFCFileDeleteView(ProjectAccessMixin, DeleteView):
    """Delete an IFC file."""

    model = IFCFile
    template_name = "environments/confirm_delete.html"

    def get_project(self):
        return self.get_object().project

    def get_success_url(self):
        messages.success(self.request, "IFC file deleted successfully.")
        return reverse("projects:ask", kwargs={"pk": self.object.project.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = f"Delete File: {self.object.name}"
        context["message"] = (
            "Are you sure you want to delete this IFC file? All extracted entities and associated chat context will be removed."
        )
        context["cancel_url"] = self.get_success_url()
        return context


class DocumentDeleteView(ProjectAccessMixin, DeleteView):
    """Delete a document."""

    model = Document
    template_name = "environments/confirm_delete.html"

    def get_project(self):
        return self.get_object().project

    def get_success_url(self):
        messages.success(self.request, "Document deleted successfully.")
        return reverse("projects:ask", kwargs={"pk": self.object.project.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = f"Delete Document: {self.object.name}"
        context["message"] = (
            "Are you sure you want to delete this document? The AI will no longer be able to reference it."
        )
        context["cancel_url"] = self.get_success_url()
        return context


class DocumentUpdateView(ProjectAccessMixin, UpdateView):
    """Replace a document."""

    model = Document
    fields = ["file"]
    template_name = "environments/file_form.html"

    def get_project(self):
        return self.get_object().project

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = f"Replace Document: {self.object.name}"
        context["help_text"] = (
            "The document name will be updated automatically to match the new file."
        )
        return context

    def form_valid(self, form):
        # 1. Save the new file content
        self.object = form.save(commit=False)
        self.object.name = self.object.file.name

        # Re-detect type
        ext = os.path.splitext(self.object.name)[1].lower()
        if ext == ".pdf":
            self.object.document_type = Document.DocumentType.PDF
        elif ext == ".docx":
            self.object.document_type = Document.DocumentType.DOCX
        elif ext == ".txt":
            self.object.document_type = Document.DocumentType.TXT

        self.object.save()

        # 2. Re-process (Extract -> Chunk -> Embed)
        # This was missing in your original code!
        processor = DocumentProcessor(self.object)
        success = processor.process()

        # 3. Store result
        self.request.session["processing_result"] = {
            "success": success,
            "filename": self.object.name,
            "type": self.object.get_document_type_display(),
            "message": "Document updated and re-indexed successfully.",
            "chunk_count": self.object.chunk_count,
            "error": self.object.error_message if not success else None,
        }

        return redirect("projects:file_processed", pk=self.object.project.pk)


class FileProcessedView(ProjectAccessMixin, TemplateView):
    """Shows the result of a file processing operation."""

    template_name = "environments/file_processed.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Retrieve results stored in session by the UpdateView
        context["project"] = self.get_project()
        context["result"] = self.request.session.pop("processing_result", None)
        return context


class FileUploadView(ProjectAccessMixin, TemplateView):
    """Unified upload page for IFC and documents"""

    template_name = "environments/file_upload.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["project"] = self.get_project()
        context["ocr_enabled"] = getattr(settings, "GLM_OCR_ENABLED", False)
        context["ocr_auto_trigger"] = getattr(settings, "GLM_OCR_AUTO_TRIGGER", False)
        return context

    def post(self, request, pk):
        """Handle file upload"""
        project = self.get_project()

        if "file" not in request.FILES:
            return JsonResponse({"error": "No file provided"}, status=400)

        uploaded_file = request.FILES["file"]
        file_ext = uploaded_file.name.lower()

        # Route to appropriate handler
        if file_ext.endswith(".ifc"):
            return self._handle_ifc_upload(project, uploaded_file)
        elif file_ext.endswith((".pdf", ".docx", ".txt")):
            return self._handle_document_upload(project, uploaded_file)
        else:
            return JsonResponse(
                {"success": False, "error": f"Unsupported file type: {file_ext}"}, status=400
            )

    def _handle_ifc_upload(self, project, uploaded_file):
        """Handle IFC file upload and parsing.

        Schema is detected before the pipeline runs. Legacy (non-IFC4) files are
        stored but not parsed; the response includes a conversion URL instead.
        """
        from ifc_processor.services.validators import sniff_schema

        try:
            if not uploaded_file.name.lower().endswith(".ifc"):
                return JsonResponse({"success": False, "error": "File must be .ifc"}, status=400)

            # Schema guard: detect before running the pipeline
            schema = sniff_schema(uploaded_file)
            is_legacy = bool(schema) and not schema.startswith("IFC4")

            ifc_file = IFCFile.objects.create(
                project=project,
                name=uploaded_file.name,
                file=uploaded_file,
                status=IFCFile.Status.PENDING,
            )

            if is_legacy:
                ifc_file.schema_version = schema
                ifc_file.status = IFCFile.Status.FAILED
                ifc_file.error_message = (
                    f"Legacy schema {schema} — conversion to IFC4 required before processing."
                )
                ifc_file.save(update_fields=["schema_version", "status", "error_message"])
                return JsonResponse(
                    {
                        "success": True,
                        "file_type": "ifc",
                        "file_name": ifc_file.name,
                        "entity_count": 0,
                        "schema_version": schema,
                        "needs_conversion": True,
                        "convert_url": reverse("projects:ifc_convert", kwargs={"pk": ifc_file.pk}),
                    }
                )

            # IFC4 — run pipeline normally
            processor = IFCProcessingService(ifc_file)
            success = processor.run_pipeline()

            if success:
                detected_schema = ifc_file.schema_version or ""
                git_hash = processor.git_commit_hash
                return JsonResponse(
                    {
                        "success": True,
                        "file_type": "ifc",
                        "file_name": ifc_file.name,
                        "entity_count": ifc_file.entity_count,
                        "schema_version": detected_schema,
                        "needs_conversion": False,
                        "git_tracked": bool(git_hash),
                        "git_commit": git_hash[:8] if git_hash else None,
                    }
                )
            else:
                return JsonResponse(
                    {
                        "success": False,
                        "error": ifc_file.error_message or "Failed to parse IFC file",
                    },
                    status=400,
                )

        except Exception as e:
            logger.exception("Error uploading IFC file: %s", e)
            return JsonResponse({"success": False, "error": str(e)}, status=500)

    def _handle_document_upload(self, project, uploaded_file):
        """Handle document (PDF/DOCX/TXT) upload"""
        try:
            name = uploaded_file.name.lower()

            # Determine document type
            if name.endswith(".pdf"):
                doc_type = Document.DocumentType.PDF
            elif name.endswith(".docx"):
                doc_type = Document.DocumentType.DOCX
            elif name.endswith(".txt"):
                doc_type = Document.DocumentType.TXT
            else:
                doc_type = Document.DocumentType.OTHER

            # Create document record
            document = Document.objects.create(
                project=project,
                name=uploaded_file.name,
                file=uploaded_file,
                document_type=doc_type,
                status=Document.Status.PENDING,
            )

            # Initialize and run the processor explicitly
            processor = DocumentProcessor(document)
            success = processor.process()

            if success:
                response_data = {
                    "success": True,
                    "file_type": doc_type,
                    "file_name": document.name,
                    "chunk_count": document.chunk_count,
                }

                # Signal the frontend to open an OCR WebSocket if auto-trigger is on.
                # The frontend drives the OCR job; this keeps analyze_document() as a
                # single dispatch point swappable for .delay() in a future Celery migration.
                from django.conf import settings

                ocr_enabled = getattr(settings, "GLM_OCR_ENABLED", False)
                auto_trigger = getattr(settings, "GLM_OCR_AUTO_TRIGGER", False)
                _pdf_types = (Document.DocumentType.PDF, Document.DocumentType.SCANNED_PDF)
                if (
                    ocr_enabled
                    and document.document_type in _pdf_types
                    and document.ocr_status == Document.OcrStatus.NONE
                ):
                    document.refresh_from_db(fields=["id"])
                    response_data["ocr_ws_url"] = (
                        f"ws/projects/{document.project_id}/documents/{document.id}/ocr/"
                    )
                    if auto_trigger and document.has_visual_content:
                        response_data["needs_ocr"] = True

                return JsonResponse(response_data)
            else:
                return JsonResponse(
                    {"success": False, "error": document.error_message or "Processing failed"},
                    status=400,
                )
            # --- END FIX ---

        except Exception as e:
            logger.exception(f"Error uploading document: {e}")
            return JsonResponse({"success": False, "error": str(e)}, status=500)


class IFCSchemaConvertView(ProjectAccessMixin, View):
    """Dedicated page to convert a legacy IFC schema (e.g. IFC2X3) to IFC4."""

    def get_project(self):
        """Resolve project via the IFC file's pk kwarg."""
        ifc_file = get_object_or_404(IFCFile, pk=self.kwargs["pk"])
        project = ifc_file.project
        if not ProjectAccessService.can_access(self.request.user, project):
            raise PermissionDenied
        return project

    def get(self, request, pk):
        ifc_file = get_object_or_404(IFCFile, pk=pk)
        project = self.get_project()
        return render(
            request,
            "environments/schema_convert.html",
            {
                "project": project,
                "ifc_file": ifc_file,
            },
        )

    def post(self, request, pk):
        from ifc_processor.services.schema_converter import IFCSchemaConverterService

        ifc_file = get_object_or_404(IFCFile, pk=pk)
        project = self.get_project()
        result = IFCSchemaConverterService(ifc_file).convert()
        return render(
            request,
            "environments/schema_convert.html",
            {
                "project": project,
                "ifc_file": ifc_file,
                "result": result,
            },
        )


class DocumentOCRView(LoginRequiredMixin, View):
    """Dedicated page to run GLM-OCR on a document."""

    def get(self, request, pk):
        """Render the OCR scan page for a document."""
        document = get_object_or_404(
            Document.objects.select_related("project"),
            pk=pk,
        )
        project = document.project
        if not ProjectAccessService.can_access(request.user, project):
            raise PermissionDenied

        if not getattr(settings, "GLM_OCR_ENABLED", False):
            messages.error(request, "OCR subsystem is disabled.")
            return redirect("projects:detail", pk=project.pk)

        return render(
            request,
            "environments/ocr_scan.html",
            {
                "document": document,
                "project": project,
            },
        )


# --- IFC Explorer ---


class ExploreView(ProjectTabMixin, TemplateView):
    """Explore tab — spatial hierarchy browser for IFC entities."""

    active_tab = "explore"
    template_name = "environments/project_detail.html"

    def get_context_data(self, **kwargs: object) -> dict:
        context = super().get_context_data(**kwargs)
        project = self.get_project()

        completed_ifc_files = project.ifc_files.filter(status=IFCFile.Status.COMPLETED).only(
            "id", "name", "entity_count"
        )

        # Resolve selected IFC file from URL kwarg or default to first
        ifc_id = self.kwargs.get("ifc_id")
        if ifc_id:
            selected_ifc = get_object_or_404(
                IFCFile, pk=ifc_id, project=project, status=IFCFile.Status.COMPLETED
            )
        else:
            selected_ifc = completed_ifc_files.first()

        root_nodes = []
        if selected_ifc:
            root_nodes = list(
                IFCSpatialElement.objects.filter(
                    ifc_file=selected_ifc,
                    parent__isnull=True,
                )
                .select_related("entity")
                .annotate(entity_count=Count("contained_entities"))
                .order_by("spatial_type", "entity__name")
            )

        context["completed_ifc_files"] = completed_ifc_files
        context["selected_ifc"] = selected_ifc
        context["root_nodes"] = root_nodes
        return context


class ExploreTreePartial(ProjectAccessMixin, View):
    """HTMX partial: returns child spatial nodes for a given parent."""

    def get(self, request, pk: str, ifc_id: str) -> object:
        project = self.get_project()
        ifc_file = get_object_or_404(
            IFCFile, pk=ifc_id, project=project, status=IFCFile.Status.COMPLETED
        )

        spatial_id = request.GET.get("spatial_id", "")

        if spatial_id:
            nodes = list(
                IFCSpatialElement.objects.filter(
                    ifc_file=ifc_file,
                    parent_id=spatial_id,
                )
                .select_related("entity")
                .annotate(entity_count=Count("contained_entities"))
                .order_by("spatial_type", "entity__name")
            )
        else:
            nodes = list(
                IFCSpatialElement.objects.filter(
                    ifc_file=ifc_file,
                    parent__isnull=True,
                )
                .select_related("entity")
                .annotate(entity_count=Count("contained_entities"))
                .order_by("spatial_type", "entity__name")
            )

        return render(
            request,
            "ifc_processor/explore/_tree_nodes.html",
            {
                "nodes": nodes,
                "ifc_file": ifc_file,
                "project": project,
            },
        )


def _get_spatial_descendant_ids(spatial_id: str, ifc_file: IFCFile) -> list:
    """Return spatial_id plus all descendant IDs (Python walk — tree is tiny)."""
    from collections import defaultdict

    all_spatial = IFCSpatialElement.objects.filter(ifc_file=ifc_file).values("id", "parent_id")
    children_map = defaultdict(list)
    for s in all_spatial:
        if s["parent_id"]:
            children_map[s["parent_id"]].append(s["id"])

    from uuid import UUID

    root = UUID(spatial_id) if isinstance(spatial_id, str) else spatial_id
    result = []
    stack = [root]
    while stack:
        current = stack.pop()
        result.append(current)
        stack.extend(children_map.get(current, []))
    return result


class ExploreEntitiesPartial(ProjectAccessMixin, View):
    """HTMX partial: paginated entity table filtered by spatial context and/or type."""

    PAGE_SIZE = 25

    def get(self, request, pk: str, ifc_id: str) -> object:
        project = self.get_project()
        ifc_file = get_object_or_404(
            IFCFile, pk=ifc_id, project=project, status=IFCFile.Status.COMPLETED
        )

        spatial_id = request.GET.get("spatial_id", "")
        ifc_type = request.GET.get("type", "")
        q = request.GET.get("q", "").strip()
        page_num = max(1, int(request.GET.get("page", 1) or 1))

        filters: dict = {"ifc_file": ifc_file}
        if spatial_id:
            descendant_ids = _get_spatial_descendant_ids(spatial_id, ifc_file)
            filters["spatial_container_id__in"] = descendant_ids
        if ifc_type:
            filters["ifc_type"] = ifc_type

        qs = (
            IFCEntity.objects.filter(**filters)
            .select_related("element_type", "spatial_container__entity")
            .only(
                "id",
                "global_id",
                "ifc_type",
                "name",
                "spatial_container",
                "element_type__name",
            )
            .order_by("ifc_type", "name")
        )
        if q:
            qs = qs.filter(Q(name__icontains=q) | Q(global_id__icontains=q))

        paginator = Paginator(qs, self.PAGE_SIZE)
        page_obj = paginator.get_page(page_num)

        # Type chips with counts (scoped to spatial context, ignoring active type filter)
        type_filter = {k: v for k, v in filters.items() if k != "ifc_type"}
        type_counts = list(
            IFCEntity.objects.filter(**type_filter)
            .values("ifc_type")
            .annotate(count=Count("id"))
            .order_by("ifc_type")
        )

        # Build breadcrumb from spatial ancestor chain
        breadcrumb = _build_spatial_breadcrumb(spatial_id) if spatial_id else []

        return render(
            request,
            "ifc_processor/explore/_entity_table.html",
            {
                "page_obj": page_obj,
                "ifc_file": ifc_file,
                "project": project,
                "type_counts": type_counts,
                "available_types": [tc["ifc_type"] for tc in type_counts],
                "active_type": ifc_type,
                "current_spatial_id": spatial_id,
                "breadcrumb": breadcrumb,
                "current_q": q,
            },
        )


def _build_spatial_breadcrumb(spatial_id: str) -> list[dict]:
    """Walk up the spatial tree to build a breadcrumb list [{name, spatial_type, id}, ...]."""
    breadcrumb = []
    try:
        node = IFCSpatialElement.objects.select_related("entity", "parent").get(pk=spatial_id)
        while node:
            breadcrumb.append(
                {
                    "name": node.entity.name or f"({node.get_spatial_type_display()})",
                    "spatial_type": node.spatial_type,
                    "id": str(node.pk),
                }
            )
            node = node.parent
        breadcrumb.reverse()
    except IFCSpatialElement.DoesNotExist:
        pass
    return breadcrumb


class ExploreEntityDetailPartial(ProjectAccessMixin, View):
    """HTMX partial: property inspector for a single IFC entity."""

    def get(self, request, pk: str, ifc_id: str, entity_id: str) -> object:
        project = self.get_project()
        ifc_file = get_object_or_404(IFCFile, pk=ifc_id, project=project)
        entity = get_object_or_404(
            IFCEntity.objects.select_related("spatial_container__entity"),
            pk=entity_id,
            ifc_file=ifc_file,
        )

        property_sets: dict = {}
        quantity_sets: dict = {}
        type_sets: dict = {}

        for key, value in (entity.properties or {}).items():
            parts = key.split(".")
            if parts[0] == "Type" and len(parts) >= 3:
                pset = parts[1]
                prop = ".".join(parts[2:])
                type_sets.setdefault(pset, {})[prop] = value
            elif parts[0].startswith("Qto_") and len(parts) >= 2:
                pset = parts[0]
                prop = ".".join(parts[1:])
                quantity_sets.setdefault(pset, {})[prop] = value
            elif len(parts) >= 2:
                pset = parts[0]
                prop = ".".join(parts[1:])
                property_sets.setdefault(pset, {})[prop] = value
            else:
                property_sets.setdefault("Other", {})[key] = value

        # Build quantity_units: {prop_name: unit_label} from project units
        # and apply unit-aware rounding to quantity values.
        quantity_units: dict[str, str] = {}
        project_units = ifc_file.project_units or {}
        if project_units:
            for props in quantity_sets.values():
                for prop_name in props:
                    unit_type = get_unit_type_for_quantity(prop_name)
                    if unit_type and unit_type in project_units:
                        unit_label = project_units[unit_type]
                        quantity_units[prop_name] = unit_label
                        val = props[prop_name]
                        if isinstance(val, float):
                            props[prop_name] = smart_round(val, unit_label)

        # Build spatial breadcrumb from entity's container chain
        spatial_breadcrumb = []
        if entity.spatial_container:
            spatial_breadcrumb = _build_spatial_breadcrumb(str(entity.spatial_container_id))

        return render(
            request,
            "ifc_processor/explore/_entity_detail.html",
            {
                "entity": entity,
                "project": project,
                "property_sets": property_sets,
                "quantity_sets": quantity_sets,
                "type_sets": type_sets,
                "quantity_units": quantity_units,
                "spatial_breadcrumb": spatial_breadcrumb,
            },
        )


class ExploreExportView(ProjectAccessMixin, View):
    """CSV export of filtered IFC entities."""

    def get(self, request, pk: str, ifc_id: str) -> HttpResponse:
        project = self.get_project()
        ifc_file = get_object_or_404(
            IFCFile, pk=ifc_id, project=project, status=IFCFile.Status.COMPLETED
        )

        spatial_id = request.GET.get("spatial_id", "")
        ifc_type = request.GET.get("type", "")
        q = request.GET.get("q", "").strip()

        filters: dict = {"ifc_file": ifc_file}
        if spatial_id:
            descendant_ids = _get_spatial_descendant_ids(spatial_id, ifc_file)
            filters["spatial_container_id__in"] = descendant_ids
        if ifc_type:
            filters["ifc_type"] = ifc_type

        qs = IFCEntity.objects.filter(**filters).order_by("ifc_type", "name")
        if q:
            qs = qs.filter(Q(name__icontains=q) | Q(global_id__icontains=q))

        response = HttpResponse(content_type="text/csv")
        safe_name = ifc_file.name.replace('"', "")
        response["Content-Disposition"] = f'attachment; filename="{safe_name}_entities.csv"'

        writer = csv.writer(response)
        writer.writerow(["GlobalID", "Type", "Name", "Element Type", "Spatial Container"])
        for entity in (
            qs.select_related("element_type", "spatial_container__entity")
            .only(
                "global_id",
                "ifc_type",
                "name",
                "spatial_container",
                "element_type__name",
            )
            .iterator()
        ):
            container_name = ""
            if entity.spatial_container and entity.spatial_container.entity:
                container_name = entity.spatial_container.entity.name or ""
            writer.writerow(
                [
                    entity.global_id,
                    entity.ifc_type,
                    entity.name or "",
                    entity.element_type.name if entity.element_type else "",
                    container_name,
                ]
            )

        return response


# --- Dynamic Generic Schedule ---


class ScheduleView(ProjectTabMixin, TemplateView):
    """Schedule tab — dynamic element schedule for a selected IFC model."""

    active_tab = "schedule"
    template_name = "environments/project_detail.html"

    def get_context_data(self, **kwargs: object) -> dict:
        context = super().get_context_data(**kwargs)
        project = self.get_project()

        completed_ifc_files = project.ifc_files.filter(status=IFCFile.Status.COMPLETED).only(
            "id", "name", "entity_count"
        )

        ifc_id = self.kwargs.get("ifc_id")
        if ifc_id:
            selected_ifc = get_object_or_404(
                IFCFile, pk=ifc_id, project=project, status=IFCFile.Status.COMPLETED
            )
        else:
            selected_ifc = completed_ifc_files.first()

        # Types pre-selected from Explore "Schedule this type" link (?types=IfcWall,…)
        preselected_types = [
            t.strip() for t in self.request.GET.get("types", "").split(",") if t.strip()
        ]

        available_types: list[dict] = []
        storeys: list[str] = []
        if selected_ifc:
            svc = GenericScheduleService(selected_ifc)
            available_types = svc.get_available_types()
            storeys = svc.get_storeys()

        context.update(
            {
                "completed_ifc_files": completed_ifc_files,
                "selected_ifc": selected_ifc,
                "available_types": available_types,
                "type_categories": IFC_TYPE_CATEGORIES,
                "preselected_types": preselected_types,
                "storeys": storeys,
            }
        )
        return context


class ScheduleTablePartial(ProjectAccessMixin, View):
    """HTMX partial — schedule tables for selected IFC types."""

    def get(self, request, pk: str, ifc_id: str) -> HttpResponse:
        project = self.get_project()
        ifc_file = get_object_or_404(
            IFCFile, pk=ifc_id, project=project, status=IFCFile.Status.COMPLETED
        )

        raw = request.GET.get("types", "")
        ifc_types = [t.strip() for t in raw.split(",") if t.strip()]
        storey = request.GET.get("storey") or None
        group_by = request.GET.get("group_by", "")

        if not ifc_types:
            return render(request, "ifc_processor/schedule/_no_selection.html", {})

        svc = GenericScheduleService(ifc_file)

        if group_by == "type":
            results = svc.get_type_grouped_schedule(ifc_types, storey=storey)
            return render(
                request,
                "ifc_processor/schedule/_type_grouped_tables.html",
                {
                    "results": results,
                    "project": project,
                    "ifc_file": ifc_file,
                    "selected_types": ifc_types,
                    "selected_storey": storey or "",
                },
            )

        results = svc.get_schedule(ifc_types, storey=storey)
        return render(
            request,
            "ifc_processor/schedule/_schedule_tables.html",
            {
                "results": results,
                "project": project,
                "ifc_file": ifc_file,
                "selected_types": ifc_types,
                "selected_storey": storey or "",
            },
        )


class ScheduleExcelExportView(ProjectAccessMixin, View):
    """Excel export — one sheet per IFC type with dynamically discovered columns."""

    def get(self, request, pk: str, ifc_id: str) -> HttpResponse:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        project = self.get_project()
        ifc_file = get_object_or_404(
            IFCFile, pk=ifc_id, project=project, status=IFCFile.Status.COMPLETED
        )

        raw = request.GET.get("types", "")
        ifc_types = [t.strip() for t in raw.split(",") if t.strip()]
        storey = request.GET.get("storey") or None
        group_by = request.GET.get("group_by", "")

        svc = GenericScheduleService(ifc_file)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default empty sheet

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1F4E79")
        center = Alignment(horizontal="center")

        used_names: set[str] = set()

        if group_by == "type":
            results = svc.get_type_grouped_schedule(ifc_types, storey=storey)
            for ifc_type, result in results.items():
                base = ifc_type.replace("Ifc", "")[:28]
                sheet_name = base
                suffix = 2
                while sheet_name in used_names:
                    sheet_name = f"{base[:25]}_{suffix}"
                    suffix += 1
                used_names.add(sheet_name)

                ws = wb.create_sheet(title=sheet_name)

                # Header: Type Name, Count, then property columns
                headers = ["Type Name", "Count"] + [
                    f"{col.label} ({col.unit})" if col.unit else col.label for col in result.columns
                ]
                ws.append(headers)
                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center

                for group in result.groups:
                    ws.append([group.element_type_name, group.count, *group.values])

                for col in ws.columns:
                    max_len = max((len(str(c.value or "")) for c in col), default=10)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
        else:
            results = svc.get_schedule(ifc_types, storey=storey)
            for ifc_type, result in results.items():
                base = ifc_type.replace("Ifc", "")[:28]
                sheet_name = base
                suffix = 2
                while sheet_name in used_names:
                    sheet_name = f"{base[:25]}_{suffix}"
                    suffix += 1
                used_names.add(sheet_name)

                ws = wb.create_sheet(title=sheet_name)

                display_headers = [
                    f"{col.label} ({col.unit})" if col.unit else col.label for col in result.columns
                ]
                ws.append(display_headers)
                for col_idx in range(1, len(display_headers) + 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center

                for row in result.rows:
                    ws.append(row)

                for col in ws.columns:
                    max_len = max((len(str(c.value or "")) for c in col), default=10)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        safe_name = ifc_file.name.replace('"', "")
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        response["Content-Disposition"] = f'attachment; filename="{safe_name}_schedule_{ts}.xlsx"'
        wb.save(response)
        return response


# --- People (members + functional roles) ---


class PeopleView(ProjectTabMixin, TemplateView):
    """Project Settings → People page.

    Two sections: Access (ProjectMembership CRUD for OWNERs) and Functional
    roles (ProjectRole list — read-only in this PR, see 7D-milestones.md).
    """

    active_tab = "people"
    template_name = "environments/people.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        project = self.get_project()
        user_permission = ProjectAccessService.user_permission(self.request.user, project)

        context["memberships"] = list(ProjectAccessService.members_ordered(project))
        context["facility_roles"] = list(
            ProjectRole.objects.filter(project=project)
            .select_related("user")
            .order_by("user__username", "role", "-valid_from")
        )
        can_admin = ProjectAccessService.can_admin(self.request.user, project)
        context["can_manage_access"] = can_admin
        # Role-CRUD authorization currently matches access-tier CRUD (OWNER-
        # only). Kept as a separate key so widening it later (e.g. to an
        # ADMIN tier) touches exactly one line here.
        context["can_manage_roles"] = can_admin
        context["user_permission"] = user_permission
        context["permission_choices"] = [
            # OWNER is excluded from the mutable choices — ownership goes via
            # the transfer modal, not a plain dropdown.
            (ProjectMembership.Permission.EDITOR, "Editor"),
            (ProjectMembership.Permission.VIEWER, "Viewer"),
        ]
        context["role_choices"] = ProjectRole.Role.choices
        return context


def _render_member_row(request, project, membership):
    return render(
        request,
        "environments/partials/_member_row.html",
        {
            "project": project,
            "membership": membership,
            "can_manage_access": True,
            "permission_choices": [
                (ProjectMembership.Permission.EDITOR, "Editor"),
                (ProjectMembership.Permission.VIEWER, "Viewer"),
            ],
        },
    )


class UserSearchView(ProjectAccessMixin, View):
    """JSON GET: typeahead over all system users.

    Returns ``{"matches": [...], "too_short": bool, "query": str, "scope": str}``
    where each match is ``{"id": int, "username": str, "email": str}``. The
    frontend (vanilla ``fetch`` in ``people.html``) renders the dropdown.

    HTML + HTMX was tried first and the response never reached the user —
    silent failure mode was too hard to diagnose. JSON lets the browser's
    Network tab show every round-trip, and the client-side render path is
    fully debuggable with console.log.
    """

    MIN_Q = 2
    LIMIT = 8
    SCOPES = ("member", "role")

    def get(self, request, pk):
        project = self.get_project()
        if not ProjectAccessService.can_admin(request.user, project):
            raise PermissionDenied

        query = (request.GET.get("q") or "").strip()
        scope = request.GET.get("scope") or "member"
        if scope not in self.SCOPES:
            scope = "member"

        if len(query) < self.MIN_Q:
            return JsonResponse({"too_short": True, "matches": [], "query": query, "scope": scope})

        user_model = get_user_model()
        qs = user_model.objects.filter(
            Q(username__icontains=query) | Q(email__icontains=query)
        ).order_by("username")

        # For the member-add picker, hide users already in the project — adding
        # them again is a no-op and crowds the list. For the role-add picker we
        # keep everyone, since a single user can hold both a membership and
        # multiple functional roles.
        if scope == "member":
            member_ids = ProjectMembership.objects.filter(project=project).values_list(
                "user_id", flat=True
            )
            qs = qs.exclude(pk__in=list(member_ids))

        matches = [{"id": u.pk, "username": u.username, "email": u.email} for u in qs[: self.LIMIT]]
        return JsonResponse(
            {"too_short": False, "matches": matches, "query": query, "scope": scope}
        )


class MemberAddView(ProjectAccessMixin, View):
    """HTMX POST: add a member by ``user_id`` picked from the typeahead.

    Email-based invite is a separate, client-only flow (``_invite_by_email.html``)
    that opens a ``mailto:`` — this endpoint expects an already-registered user.
    """

    def post(self, request, pk):
        project = self.get_project()
        if not ProjectAccessService.can_admin(request.user, project):
            raise PermissionDenied

        permission = request.POST.get("permission") or ProjectMembership.Permission.VIEWER
        user_id = (request.POST.get("user_id") or "").strip()

        if not user_id:
            return HttpResponse("Pick a user from the suggestions first.", status=400)

        user_model = get_user_model()
        user = user_model.objects.filter(pk=user_id).first()
        if user is None:
            return HttpResponse("User not found.", status=404)

        if user.pk == project.owner_id and permission != ProjectMembership.Permission.OWNER:
            return HttpResponse("The project owner already has OWNER membership.", status=400)

        try:
            membership = ProjectAccessService.add_member(
                project=project,
                user=user,
                permission=permission,
                invited_by=request.user,
            )
        except ProjectAccessError as exc:
            return HttpResponse(str(exc), status=400)

        return _render_member_row(request, project, membership)


class MemberChangePermissionView(ProjectAccessMixin, View):
    """HTMX POST: change an existing member's permission."""

    def post(self, request, pk, user_id):
        project = self.get_project()
        if not ProjectAccessService.can_admin(request.user, project):
            raise PermissionDenied

        new_permission = request.POST.get("permission")
        if new_permission not in ProjectMembership.Permission.values:
            return HttpResponse("Unknown permission.", status=400)

        user_model = get_user_model()
        target = get_object_or_404(user_model, pk=user_id)

        try:
            membership = ProjectAccessService.change_permission(
                project=project,
                user=target,
                new_permission=new_permission,
            )
        except OwnerDemotionBlocked as exc:
            return HttpResponse(str(exc), status=400)
        except ProjectAccessError as exc:
            return HttpResponse(str(exc), status=400)
        except ProjectMembership.DoesNotExist:
            return HttpResponse("User is not a member.", status=404)

        return _render_member_row(request, project, membership)


class MemberRemoveView(ProjectAccessMixin, View):
    """HTMX POST/DELETE: remove a member from the project."""

    def post(self, request, pk, user_id):
        return self._remove(request, pk, user_id)

    def delete(self, request, pk, user_id):
        return self._remove(request, pk, user_id)

    def _remove(self, request, pk, user_id):
        project = self.get_project()
        if not ProjectAccessService.can_admin(request.user, project):
            raise PermissionDenied

        user_model = get_user_model()
        target = get_object_or_404(user_model, pk=user_id)

        try:
            ProjectAccessService.remove_member(project=project, user=target)
        except LastOwnerRemovalBlocked as exc:
            return HttpResponse(str(exc), status=400)
        except ProjectMembership.DoesNotExist:
            return HttpResponse("User is not a member.", status=404)

        # HTMX swap strategy: return empty 200 and let the row hx-swap="outerHTML"
        # on the client remove the <tr>.
        return HttpResponse("", status=200)


class TransferOwnershipView(ProjectAccessMixin, View):
    """HTMX POST: transfer ownership to another existing member."""

    def post(self, request, pk):
        project = self.get_project()
        if not ProjectAccessService.can_delete(request.user, project):
            raise PermissionDenied

        new_owner_id = request.POST.get("new_owner_id")
        if not new_owner_id:
            return HttpResponse("new_owner_id is required.", status=400)

        user_model = get_user_model()
        new_owner = get_object_or_404(user_model, pk=new_owner_id)

        if not ProjectAccessService.can_access(new_owner, project):
            return HttpResponse(
                "Target user must already be a member before receiving ownership.",
                status=400,
            )

        ProjectAccessService.transfer_ownership(project=project, new_owner=new_owner)
        messages.success(request, f"Ownership transferred to {new_owner.username}.")
        # Full refresh — ownership change reshuffles every row's control set.
        return HttpResponse(
            status=204, headers={"HX-Redirect": reverse("projects:people", args=[project.pk])}
        )


# --- Functional-role (ProjectRole) CRUD --------------------------------------
# These live next to the access-tier views so the entire People-page surface
# stays in one file. The role service is imported lazily inside each handler to
# avoid a hard cycle between environments and facilities at module-load time.


def _render_role_row(request, project, role):
    return render(
        request,
        "environments/partials/_role_row.html",
        {
            "project": project,
            "role": role,
            "can_manage_roles": True,
        },
    )


def _parse_date_or_none(value: str | None):
    """Parse a 'YYYY-MM-DD' form field into an aware datetime at 00:00 local.

    Returns None for empty/whitespace input. Raises ValueError on bad format.
    """
    from django.utils import timezone as dj_timezone

    if not value or not value.strip():
        return None
    parsed = datetime.strptime(value.strip(), "%Y-%m-%d")
    return dj_timezone.make_aware(parsed)


class RoleAddView(ProjectAccessMixin, View):
    """HTMX POST: assign a functional role (ProjectRole) to any user."""

    def post(self, request, pk):
        from facilities.services import ProjectRoleError, ProjectRoleService

        project = self.get_project()
        if not ProjectAccessService.can_admin(request.user, project):
            raise PermissionDenied

        user_id = (request.POST.get("user_id") or "").strip()
        role_value = (request.POST.get("role") or "").strip()
        valid_from_raw = request.POST.get("valid_from")
        valid_until_raw = request.POST.get("valid_until")

        if not user_id:
            return HttpResponse("Pick a user from the suggestions.", status=400)
        if role_value not in ProjectRole.Role.values:
            return HttpResponse("Pick a functional role.", status=400)

        user_model = get_user_model()
        target = user_model.objects.filter(pk=user_id).first()
        if target is None:
            return HttpResponse("User not found.", status=404)

        try:
            valid_from = _parse_date_or_none(valid_from_raw)
            valid_until = _parse_date_or_none(valid_until_raw)
        except ValueError:
            return HttpResponse("Dates must be YYYY-MM-DD.", status=400)

        try:
            role = ProjectRoleService.add_role(
                project=project,
                user=target,
                role=role_value,
                valid_from=valid_from,
                valid_until=valid_until,
            )
        except ProjectRoleError as exc:
            return HttpResponse(str(exc), status=400)

        # Re-select with user for template display without another query.
        role = ProjectRole.objects.select_related("user").get(pk=role.pk)
        return _render_role_row(request, project, role)


class RoleRemoveView(ProjectAccessMixin, View):
    """HTMX POST/DELETE: remove a functional-role assignment."""

    def post(self, request, pk, role_id):
        return self._remove(request, pk, role_id)

    def delete(self, request, pk, role_id):
        return self._remove(request, pk, role_id)

    def _remove(self, request, pk, role_id):
        from facilities.services import ProjectRoleError, ProjectRoleService

        project = self.get_project()
        if not ProjectAccessService.can_admin(request.user, project):
            raise PermissionDenied

        try:
            ProjectRoleService.remove_role(project=project, role_id=role_id)
        except ProjectRoleError as exc:
            return HttpResponse(str(exc), status=404)

        # Empty body + 200 — the <tr>'s hx-swap="outerHTML" strips the row.
        return HttpResponse("", status=200)


class ScheduleExportView(ProjectAccessMixin, View):
    """CSV export of the door/window schedule for a selected IFC model."""

    def get(self, request, pk: str, ifc_id: str) -> HttpResponse:
        project = self.get_project()
        ifc_file = get_object_or_404(
            IFCFile, pk=ifc_id, project=project, status=IFCFile.Status.COMPLETED
        )

        element_type = request.GET.get("type", "all")
        selected_storey = request.GET.get("storey", "")

        svc = DoorWindowScheduleService(ifc_file)
        rows: list[dict] = []
        if element_type in ("all", "doors"):
            for r in svc.get_doors(storey=selected_storey):
                r["element_type"] = "Door"
                rows.append(r)
        if element_type in ("all", "windows"):
            for r in svc.get_windows(storey=selected_storey):
                r["element_type"] = "Window"
                rows.append(r)

        response = HttpResponse(content_type="text/csv")
        safe_name = ifc_file.name.replace('"', "")
        response["Content-Disposition"] = (
            f'attachment; filename="{safe_name}_door_window_schedule.csv"'
        )

        writer = csv.writer(response)
        writer.writerow(
            [
                "Type",
                "Mark",
                "Level",
                "Room",
                "Width (m)",
                "Height (m)",
                "Fire Rating",
                "External",
                "Acoustic Rating",
                "Security Rating",
                "Accessible",
                "U-Value (W/m²K)",
                "Reference",
                "GlobalID",
            ]
        )
        for r in rows:
            writer.writerow(
                [
                    r.get("element_type", ""),
                    r.get("mark", ""),
                    r.get("level", ""),
                    r.get("room", ""),
                    r.get("width", ""),
                    r.get("height", ""),
                    r.get("fire_rating", ""),
                    r.get("is_external", ""),
                    r.get("acoustic_rating", ""),
                    r.get("security_rating", ""),
                    r.get("handicap_accessible", ""),
                    r.get("thermal_transmittance", ""),
                    r.get("reference", ""),
                    r.get("global_id", ""),
                ]
            )

        return response
