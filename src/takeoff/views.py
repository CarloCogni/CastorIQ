# takeoff/views.py
"""HTTP views for the Quantity Take-Off (QTO) tab.

The 5D bridge view: aggregates IFC quantities by type / level / material and
exposes unit-cost editing so EVM can read cost baselines.
"""

from __future__ import annotations

import io
import logging

from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils.decorators import method_decorator
from django.views import View
from django.views.decorators.http import require_POST
from django.views.generic import TemplateView

from core.http import toast_response, trigger_toast
from core.mixins import ProjectAccessMixin, ProjectTabMixin

from .models import QTOCache
from .services.link_analysis import LinkAnalysisService
from .services.model_inventory import ModelInventoryService
from .services.model_quantities import ModelQuantitiesService
from .services.quantity_preparation_ui import (
    build_preparation_ui,
    parse_basis_overrides_from_query,
    parse_schema_includes_from_query,
    parse_source_mappings_from_query,
)

logger = logging.getLogger(__name__)

_LINK_ANALYSIS_SESSION_KEY = "link_analysis_last_diagnostic_run"


class ModelInventoryView(ProjectTabMixin, TemplateView):
    """4D Link Analysis — schedule task ↔ model element link diagnostics (Model hub)."""

    active_tab = "castor"

    def get_context_data(self, **kwargs: object) -> dict:
        ctx = super().get_context_data(**kwargs)
        ctx["castor_subtab"] = "model_inventory"
        project = ctx["project"]
        last_run = self.request.session.get(_LINK_ANALYSIS_SESSION_KEY)
        try:
            task_page = int(self.request.GET.get("task_page") or 1)
        except (TypeError, ValueError):
            task_page = 1
        try:
            element_page = int(self.request.GET.get("element_page") or 1)
        except (TypeError, ValueError):
            element_page = 1
        analysis = LinkAnalysisService(project).build(
            task_page=task_page,
            element_page=element_page,
            search=self.request.GET.get("q") or "",
            last_diagnostic_run=last_run,
        )
        ctx["analysis"] = analysis
        # Legacy alias kept for any template that still expects inventory.
        ctx["inventory"] = analysis
        ctx["viewer_url"] = reverse("ifc_viewer:viewer", kwargs={"pk": project.pk})
        schedule_url = reverse("scheduling:schedule", kwargs={"pk": project.pk})
        ctx["apply_url"] = f"{schedule_url}?tab=fourD_link"
        ctx["time_view_url"] = f"{schedule_url}?tab=lookahead"
        ctx["schedule_url"] = f"{schedule_url}?tab=data_sources"
        ctx["quantities_url"] = reverse("takeoff:qto", kwargs={"pk": project.pk})
        ctx["entities_url"] = reverse("takeoff:model_inventory_entities", kwargs={"pk": project.pk})
        ctx["refresh_url"] = reverse("takeoff:link_analysis_refresh", kwargs={"pk": project.pk})
        return ctx


@method_decorator(require_POST, name="dispatch")
class LinkAnalysisRefreshView(ProjectAccessMixin, View):
    """Refresh analysis aggregates only — no apply/approve/unlink mutations."""

    def post(self, request, pk):  # type: ignore[override]
        project = self.get_project()
        result = LinkAnalysisService(project).run_diagnostics()
        request.session[_LINK_ANALYSIS_SESSION_KEY] = result["last_run"]
        response = redirect("takeoff:model_inventory", pk=project.pk)
        kpis = result.get("kpis") or {}
        checked = int(kpis.get("tasks_total") or result.get("tasks_total") or 0)
        linked = int(kpis.get("linked_tasks") or 0)
        return trigger_toast(
            response,
            f"Analysis refreshed — {checked} tasks checked · {linked} linked",
            level="success",
        )


class ModelInventoryEntitiesView(ProjectTabMixin, TemplateView):
    """IFC Elements list — HTMX partial on Model page; full shell for browser GET."""

    active_tab = "castor"

    def get(self, request, *args, **kwargs):  # type: ignore[override]
        project = self.get_project()
        svc = ModelInventoryService(project)
        self._entities_result = svc.list_entities(
            ifc_class=request.GET.get("ifc_class"),
            level=request.GET.get("level"),
            linked_status=request.GET.get("linked_status"),
            has_qto=request.GET.get("has_qto"),
            page=request.GET.get("page"),
            page_size=request.GET.get("page_size"),
        )
        self._entities_url = reverse("takeoff:model_inventory_entities", kwargs={"pk": project.pk})
        if request.headers.get("HX-Request"):
            return render(
                request,
                "takeoff/components/model_inventory_entities.html",
                {
                    "project": project,
                    "entities": self._entities_result,
                    "entities_url": self._entities_url,
                },
            )
        inventory = svc.build()
        self._filter_options = inventory.get("filter_options") or {
            "ifc_classes": [],
            "levels": [],
        }
        self._inventory_source_name = inventory.get("ifc_file_name") or ""
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs: object) -> dict:
        ctx = super().get_context_data(**kwargs)
        ctx["castor_subtab"] = "model_inventory_entities"
        project = ctx["project"]
        svc = ModelInventoryService(project)
        if getattr(self, "_entities_result", None) is None:
            self._entities_result = svc.list_entities(
                ifc_class=self.request.GET.get("ifc_class"),
                level=self.request.GET.get("level"),
                linked_status=self.request.GET.get("linked_status"),
                has_qto=self.request.GET.get("has_qto"),
                page=self.request.GET.get("page"),
                page_size=self.request.GET.get("page_size"),
            )
            self._entities_url = reverse(
                "takeoff:model_inventory_entities", kwargs={"pk": project.pk}
            )
        if getattr(self, "_filter_options", None) is None:
            inventory = svc.build()
            self._filter_options = inventory.get("filter_options") or {
                "ifc_classes": [],
                "levels": [],
            }
            self._inventory_source_name = inventory.get("ifc_file_name") or ""
        ctx["entities"] = self._entities_result
        ctx["entities_url"] = getattr(self, "_entities_url", None) or reverse(
            "takeoff:model_inventory_entities", kwargs={"pk": project.pk}
        )
        ctx["model_inventory_url"] = reverse("takeoff:model_inventory", kwargs={"pk": project.pk})
        ctx["filter_options"] = self._filter_options
        ctx["inventory_source_name"] = getattr(self, "_inventory_source_name", "") or ""
        return ctx


class QTOView(ProjectTabMixin, TemplateView):
    """Quantities tab — builder-led quantity preparation + model quantity reference."""

    active_tab = "castor"

    def get_context_data(self, **kwargs: object) -> dict:
        ctx = super().get_context_data(**kwargs)
        ctx["castor_subtab"] = "qto"
        project = ctx["project"]
        # Slice 3a/3c-1/3c-2a: basis_*, field_*, source_* GET params are session/UI only.
        quantities = ModelQuantitiesService(project).build()
        ctx["quantities"] = quantities
        basis_overrides = parse_basis_overrides_from_query(self.request.GET)
        schema_includes = parse_schema_includes_from_query(self.request.GET)
        source_mappings = parse_source_mappings_from_query(self.request.GET)
        ctx["qty_prep"] = build_preparation_ui(
            quantities,
            basis_overrides=basis_overrides,
            schema_includes=schema_includes,
            source_mappings=source_mappings,
        )
        ctx["missing_qto_entities_url"] = (
            reverse("takeoff:model_inventory_entities", kwargs={"pk": project.pk}) + "?has_qto=no"
        )
        ctx["model_inventory_url"] = reverse("takeoff:model_inventory", kwargs={"pk": project.pk})
        ctx["entities_url"] = reverse("takeoff:model_inventory_entities", kwargs={"pk": project.pk})
        # Presentation-only flags from existing summary rows (no re-aggregation).
        class_rows = quantities.get("by_ifc_class") or []
        ctx["quantity_classes_with_qto"] = sum(
            1 for row in class_rows if (row.get("has_ifc_qto") or 0) > 0
        )
        ctx["quantity_measure_families"] = {
            "volume": any(
                row.get("net_volume") is not None or row.get("gross_volume") is not None
                for row in class_rows
            ),
            "area": any(
                row.get("net_area") is not None or row.get("net_side_area") is not None
                for row in class_rows
            ),
            "length": any(row.get("length") is not None for row in class_rows),
        }
        # Legacy cache kept only for demoted optional tooling on main.
        ctx["qto_cache"] = QTOCache.objects.filter(project=project).first()
        return ctx


class QTODataView(ProjectAccessMixin, View):
    """JSON endpoint — QTO cache payload (excludes items_json)."""

    def get(self, request, **kwargs: object) -> JsonResponse:
        project = self.get_project()
        cache = QTOCache.objects.filter(project=project).first()
        if not cache:
            return JsonResponse({"has_data": False})
        return JsonResponse(
            {
                "has_data": True,
                "total_entities": cache.total_entities,
                "entities_with_qty": cache.entities_with_qty,
                "coverage_pct": cache.coverage_pct,
                "total_cost_estimate": cache.total_cost_estimate,
                "summary": cache.summary_json,
                "by_level": cache.by_level_json,
                "by_material": cache.by_material_json,
            }
        )


class QTORecomputeView(ProjectAccessMixin, View):
    """HTMX POST — trigger QTO recomputation and return a toast."""

    def post(self, request, **kwargs: object) -> HttpResponse:
        project = self.get_project()
        try:
            from .services.quantities import compute_qto

            result = compute_qto(project)
        except Exception as exc:
            logger.error("QTO recompute failed for project %s: %s", project.pk, exc)
            return toast_response(f"Recompute failed: {exc}", "error", status=500)

        if not result.get("has_data"):
            return toast_response("No processed IFC file found.", "error", status=404)

        return toast_response(
            f"QTO recomputed — {result['total_entities']} elements, "
            f"{result['coverage_pct']:.0f}% coverage.",
            "success",
        )


class QTOUnitCostUpdateView(ProjectAccessMixin, View):
    """HTMX POST — set unit cost for one IFC type; persist to QTOCache."""

    def post(self, request, **kwargs: object) -> HttpResponse:
        project = self.get_project()
        ifc_type = request.POST.get("ifc_type", "").strip()
        cost_raw = request.POST.get("unit_cost", "").strip()

        if not ifc_type:
            return toast_response("IFC type is required.", "error", status=400)

        cache = QTOCache.objects.filter(project=project).first()
        if not cache:
            return toast_response("No QTO data found — recompute first.", "error", status=404)

        costs = dict(cache.unit_costs_json)
        if cost_raw:
            try:
                costs[ifc_type] = float(cost_raw)
            except ValueError:
                return toast_response("Unit cost must be a number.", "error", status=400)
        else:
            costs.pop(ifc_type, None)

        cache.unit_costs_json = costs
        cache.save(update_fields=["unit_costs_json"])
        return toast_response(f"Unit cost for {ifc_type} updated.", "success")


class QTOExportView(ProjectAccessMixin, View):
    """GET — export QTO data to Excel (3 sheets: Summary / By Level / Detail)."""

    def get(self, request, **kwargs: object) -> HttpResponse:
        from openpyxl import Workbook

        project = self.get_project()
        cache = QTOCache.objects.filter(project=project).first()
        if not cache:
            return HttpResponse("No QTO data found — run Recompute first.", status=404)

        wb = Workbook()

        ws1 = wb.active
        ws1.title = "Summary"
        ws1.append(
            ["IFC Type", "Count", "Total Qty", "Unit", "Coverage %", "Unit Cost", "Total Cost"]
        )
        for row in cache.summary_json:
            ws1.append(
                [
                    row.get("type"),
                    row.get("count"),
                    row.get("total_qty"),
                    row.get("unit"),
                    row.get("coverage_pct"),
                    row.get("unit_cost"),
                    row.get("total_cost"),
                ]
            )

        ws2 = wb.create_sheet("By Level")
        ws2.append(["Level", "Entity Count", "Estimated Cost"])
        for row in cache.by_level_json:
            ws2.append([row.get("level"), row.get("entity_count"), row.get("cost")])

        ws3 = wb.create_sheet("Detail")
        ws3.append(
            [
                "Global ID",
                "Name",
                "IFC Type",
                "Level",
                "Material",
                "Quantity",
                "Unit",
                "Source",
                "Unit Cost",
                "Total Cost",
            ]
        )
        for item in cache.items_json:
            ws3.append(
                [
                    item.get("global_id"),
                    item.get("name"),
                    item.get("type"),
                    item.get("level"),
                    item.get("material"),
                    item.get("quantity"),
                    item.get("unit"),
                    item.get("source"),
                    item.get("unit_cost"),
                    item.get("total_cost"),
                ]
            )

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        safe_name = "".join(c if c.isalnum() or c in "-_ " else "_" for c in project.name)
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="qto_{safe_name}.xlsx"'
        return response
